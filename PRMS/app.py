from flask import Flask, render_template, request, redirect, url_for
import openpyxl
import os
from datetime import datetime

app = Flask(__name__)

EXCEL_FILE = 'patients.xlsx'

if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([
        "RefID", "Name", "Mobile", "Referred", "RefType", "DrName", "Status", "Sponsor",
        "CreatedBy", "Comment", "CreatedAt"
    ])
    wb.save(EXCEL_FILE)

def read_patients():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    patients = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        while len(row) < 11:
            row.append("")
        patients.append({
            "id": row[0],
            "name": row[1],
            "mobile": row[2],
            "referred": row[3],
            "reftype": row[4],
            "drname": row[5],
            "status": row[6],
            "sponsor": row[7],
            "created_by": row[8],
            "comment": row[9],
            "created_at": row[10]
        })
    return patients

def write_patient(patient):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    new_id_num = ws.max_row - 1
    ref_id = f"Ref{new_id_num+1:03d}"
    ws.append([
        ref_id, patient["name"], patient["mobile"], patient["referred"], patient["reftype"], patient["drname"],
        patient["status"], patient["sponsor"], patient["created_by"], patient["comment"], patient["created_at"]
    ])
    wb.save(EXCEL_FILE)
    return ref_id

def update_patient_status(ref_id, status):
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if row[0].value == ref_id:
            row[6].value = status
            break
    wb.save(EXCEL_FILE)

@app.route("/", methods=["GET"])
def dashboard():
    patients = read_patients()
    status_counts = {}
    for s in ["New", "On-going", "Cleared", "No-show", "Verified"]:
        status_counts[s] = sum(1 for p in patients if p["status"] == s)
    return render_template("dashboard.html", status_counts=status_counts)

@app.route("/records", methods=["GET", "POST"])
def patient_records():
    search_query = request.form.get("search") if request.method == "POST" else ""
    patients = read_patients()
    if search_query:
        patients = [p for p in patients if search_query.lower() in p["name"].lower() or search_query in str(p["mobile"])]
    return render_template("records.html", patients=patients, search_query=search_query)

@app.route("/add", methods=["GET", "POST"])
def add_patient():
    if request.method == "POST":
        data = request.form
        name = data["name"].strip()
        mobile = data["mobile"].strip()
        referred = data["referred"]
        reftype = data.get("reftype", "")
        drname = data.get("drname", "")
        status = data["status"]
        sponsor = data["sponsor"]
        created_by = data["created_by"]
        comment = data.get("comment", "")
        created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        if not name or not mobile.isdigit() or len(mobile) < 10:
            return render_template("add_patient.html", error="Name and valid mobile no. are mandatory.")
        if referred == "Doctor" and not drname:
            return render_template("add_patient.html", error="Doctor name required for referrals.")
        patient = {
            "name": name, "mobile": mobile, "referred": referred, "reftype": reftype, "drname": drname,
            "status": status, "sponsor": sponsor, "created_by": created_by, "comment": comment, "created_at": created_at
        }
        write_patient(patient)
        return redirect(url_for("dashboard"))
    return render_template("add_patient.html")

@app.route("/edit/<ref_id>", methods=["POST"])
def edit_status(ref_id):
    status = request.form["status"]
    update_patient_status(ref_id, status)
    return redirect(url_for("patient_records"))

if __name__ == "__main__":
    app.run(debug=True)